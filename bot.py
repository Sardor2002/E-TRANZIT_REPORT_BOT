import asyncio
import logging
import sqlite3
import datetime
import re
import os
import calendar
import io
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.cm as cm
import matplotlib.dates as mdates
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
import pdfplumber
from aiogram import Bot, Dispatcher, types, F
from aiogram.types import (
    ReplyKeyboardMarkup, KeyboardButton,
    InlineKeyboardMarkup, InlineKeyboardButton, CallbackQuery, FSInputFile
)

# ================= CONFIG =================
API_TOKEN = "8566622439:AAGHYrOWz47Tf2xSu6tqLCHseT1p11FTsEA"
BOT_USERNAME = "E_Tranzit_Report_Bot"  # ← O'zingizning bot username'ingizni shu yerga yozing (masalan @E_Tranzit_Report_Bot)
SUPER_ADMIN_ID = 7764313855

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)

bot = Bot(token=API_TOKEN)  # Render uchun proxy kerak emas
dp = Dispatcher()

# ================= DATABASE =================
conn = sqlite3.connect("bot_db.sqlite3", check_same_thread=False)
cursor = conn.cursor()

def init_db():
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        telegram_id INTEGER UNIQUE,
        username TEXT,
        full_name TEXT,
        role TEXT DEFAULT 'user'
    )""")
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS files (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        declaration_number TEXT UNIQUE,
        bojxona_sum INTEGER,
        deklarant_fish TEXT,
        avto TEXT,
        tirkama TEXT,
        kirish_posti TEXT,
        timestamp TEXT,
        date_only DATE
    )""")
    try:
        cursor.execute("ALTER TABLE files ADD COLUMN date_only DATE")
    except sqlite3.OperationalError:
        pass
    conn.commit()

init_db()
user_sessions = {}

# ================= KEYBOARDS =================
def main_admin_keyboard():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="👥 Foydalanuvchilar boshqaruvi")],
            [KeyboardButton(text="📍 Olot"), KeyboardButton(text="📍 Najimov"), KeyboardButton(text="📍 Jartepa")],
            [KeyboardButton(text="📍 Sirdaryo"), KeyboardButton(text="📍 Toshkent sh"), KeyboardButton(text="📍 Andarxon")],
            [KeyboardButton(text="📊 Umumiy statistika"), KeyboardButton(text="📥 Excel Hisobot")]
        ],
        resize_keyboard=True,
        input_field_placeholder="Bo'limni tanlang..."
    )

def sub_admin_keyboard():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="📊 Mening statistikam")]
        ],
        resize_keyboard=True,
        input_field_placeholder="PDF yuboring yoki statistika ko'ring"
    )

def users_list_inline():
    cursor.execute("SELECT telegram_id, username, full_name, role FROM users WHERE telegram_id != ?", (SUPER_ADMIN_ID,))
    users = cursor.fetchall()
    kb = InlineKeyboardMarkup(inline_keyboard=[])
    for tg_id, user, name, role in users:
        display_name = name if name else (f"@{user}" if user else str(tg_id))
        role_status = "👑" if role == "chief_admin" else ("⭐" if role == "sub_admin" else "👤")
        kb.inline_keyboard.append([
            InlineKeyboardButton(text="🆔", callback_data=f"show_id:{tg_id}"),
            InlineKeyboardButton(text=f"{role_status} {display_name[:15]}", callback_data=f"u_info:{tg_id}"),
            InlineKeyboardButton(text="🗑️", callback_data=f"confirm_u_del:{tg_id}")
        ])
    kb.inline_keyboard.append([InlineKeyboardButton(text="⬅️ Asosiy menyuga qaytish", callback_data="back_to_main")])
    return kb

def get_calendar_markup(year, month, target_type, target_id, side="start"):
    markup = []
    month_names = ["Yanvar", "Fevral", "Mart", "Aprel", "May", "Iyun", "Iyul", "Avgust", "Sentyabr", "Oktyabr", "Noyabr", "Dekabr"]
    markup.append([InlineKeyboardButton(text=f"📅 {month_names[month-1]} {year}", callback_data="ignore")])
    markup.append([InlineKeyboardButton(text=d, callback_data="ignore") for d in ["Du", "Se", "Ch", "Pa", "Ju", "Sh", "Ya"]])
    month_calendar = calendar.monthcalendar(year, month)
    for week in month_calendar:
        row = []
        for day in week:
            if day == 0:
                row.append(InlineKeyboardButton(text=" ", callback_data="ignore"))
            else:
                row.append(InlineKeyboardButton(text=str(day), callback_data=f"cal:{target_type}:{target_id}:{side}:{year}:{month}:{day}"))
        markup.append(row)
    markup.append([
        InlineKeyboardButton(text="⬅️", callback_data=f"nav:{target_type}:{target_id}:{side}:{year}:{month}:prev"),
        InlineKeyboardButton(text="❌ Bekor qilish", callback_data="back_to_main"),
        InlineKeyboardButton(text="➡️", callback_data=f"nav:{target_type}:{target_id}:{side}:{year}:{month}:next")
    ])
    return InlineKeyboardMarkup(inline_keyboard=markup)

# ================= PDF & TOOLS =================
def format_currency(amount):
    return f"{amount or 0:,}".replace(",", " ") + " so'm"

def extract_bojxona_service_sum(full_text):
    text = re.sub(r"\s+", " ", full_text)
    jami_match = re.findall(r"(\d{1,3}(?:\s\d{3})*),\d{2}", text)
    if not jami_match: return 0
    jami_sum = int(jami_match[-1].replace(" ", ""))
    other_services_sum = 0
    service_lines = re.findall(r"(?!Bojxona\s+servis\s+xizmatlari)([A-Za-zА-Яа-яЁёЎҚҒҲўқғҳ\s]+?)\s+(\d{1,3}(?:\s\d{3})*),\d{2}", text, re.IGNORECASE)
    for _, amount in service_lines:
        other_services_sum += int(amount.replace(" ", ""))
    bojxona_sum = jami_sum - other_services_sum
    return bojxona_sum if bojxona_sum >= 0 else 0

def parse_pdf_data(full_text):
    deklarant = "Noma'lum"
    m = re.search(r"Deklarant:\s*([A-ZА-ЯЁЎҚҒҲ]{2,})\s+([A-ZА-ЯЁЎҚҒҲ]{2,})", full_text)
    if m: deklarant = f"{m.group(1)} {m.group(2)}"
    avto, tirkama = "Yo'q", "Yo'q"
    m = re.search(r"№\s*avto/tirkama:\s*([A-Z0-9]+)\s*/\s*([A-Z0-9]+)", full_text, re.IGNORECASE)
    if m: avto, tirkama = m.group(1), m.group(2)
    kirish_posti = "Noma'lum"
    m = re.search(r"Kirish bojxona posti:\s*(.+?)(?:\s+Tashuvchi nomi:|$)", full_text, re.IGNORECASE)
    if m: kirish_posti = m.group(1).strip()
    at_number = "Topilmadi"
    m = re.search(r"(AT\d{5,})", full_text, re.IGNORECASE)
    if m: at_number = m.group(1)
    return at_number, deklarant, avto, tirkama, kirish_posti

def get_user_role(user_id):
    if user_id == SUPER_ADMIN_ID: return "super_admin"
    cursor.execute("SELECT role FROM users WHERE telegram_id=?", (user_id,))
    r = cursor.fetchone()
    role = r[0] if r else None
    if role == "chief_admin":
        return "chief_admin"
    return role

def is_admin(role):
    return role in ["super_admin", "chief_admin"]

# ================= HANDLERS =================
@dp.message(F.text == "/start")
async def cmd_start(message: types.Message):
    user_id = message.from_user.id
    username = message.from_user.username or "yo'q"
    full_name = message.from_user.full_name or "Noma'lum"

    cursor.execute("""
        INSERT INTO users (telegram_id, username, full_name, role)
        VALUES (?, ?, ?, ?)
        ON CONFLICT(telegram_id) DO UPDATE SET
        username=excluded.username,
        full_name=excluded.full_name
    """, (user_id, username, full_name, 'user'))
    conn.commit()

    role = get_user_role(user_id)

    # Yangi foydalanuvchi start bosganini bildirish (faqat yangi user uchun)
    if role == 'user':
        notify_text = f"🆕 Yangi foydalanuvchi start bosdi!\n👤 {full_name}\nID: {user_id}\n@{username}"
        try:
            await bot.send_message(SUPER_ADMIN_ID, notify_text)
        except Exception as e:
            logging.error(f"Super admin ga xabar yuborish xatosi: {e}")
        cursor.execute("SELECT telegram_id FROM users WHERE role = 'chief_admin'")
        chiefs = cursor.fetchall()
        for (chief_id,) in chiefs:
            try:
                await bot.send_message(chief_id, notify_text)
            except Exception as e:
                logging.error(f"Chief admin ({chief_id}) ga xabar yuborish xatosi: {e}")

    if is_admin(role):
        await message.answer("🏢 **Boshqaruv paneli**", reply_markup=main_admin_keyboard(), parse_mode="Markdown")
    elif role == "sub_admin":
        await message.answer("📝 **Xush kelibsiz!**\nPDF fayllarni yuborishingiz mumkin.", reply_markup=sub_admin_keyboard(), parse_mode="Markdown")
    else:
        await message.answer("⏳ **Kirish huquqi kutilmoqda...**\nAdmin tasdiqlashini kiting.")

@dp.message(F.document)
async def handle_pdf(message: types.Message):
    role = get_user_role(message.from_user.id)
    if role is None:
        await message.reply(f"Botga shaxsiy chatda start bosing: t.me/{BOT_USERNAME}")
        return
    if role not in ["sub_admin", "chief_admin", "super_admin"]: return
    if not message.document.file_name.lower().endswith(".pdf"): return

    msg = await bot.send_message(message.from_user.id, "🔄 **Tahlil qilinmoqda...**", parse_mode="Markdown")
    path = f"tmp_{message.document.file_unique_id}.pdf"

    try:
        file = await bot.get_file(message.document.file_id)
        await bot.download_file(file.file_path, path)
        full_text = ""
        with pdfplumber.open(path) as pdf:
            for p in pdf.pages: full_text += (p.extract_text() or "") + "\n"
        at, dekl, avto, tirk, post = parse_pdf_data(full_text)
        summa = extract_bojxona_service_sum(full_text)

        if at == "Topilmadi":
            await bot.edit_message_text("❌ **Xatolik:** PDF ichidan AT raqami topilmadi.", chat_id=message.from_user.id, message_id=msg.message_id)
            return

        cursor.execute("""INSERT INTO files
            (user_id, declaration_number, bojxona_sum, deklarant_fish, avto, tirkama, kirish_posti, timestamp, date_only)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (message.from_user.id, at, summa, dekl, avto, tirk, post,
             datetime.datetime.now().strftime("%Y-%m-%d %H:%M"), datetime.date.today().isoformat()))
        conn.commit()

        result_text = (
            f"✅ **Muvaffaqiyatli qabul qilindi!**\n\n"
            f"📄 **AT:** `{at}`\n👤 **Deklarant:** {dekl}\n"
            f"🚛 **Avto:** {avto} / {tirk}\n🏢 **Post:** {post}\n💰 **Summa:** {format_currency(summa)}"
        )
        await bot.edit_message_text(result_text, chat_id=message.from_user.id, message_id=msg.message_id, parse_mode="Markdown")

    except sqlite3.IntegrityError:
        cursor.execute("""
            SELECT users.full_name, users.username, users.telegram_id
            FROM files JOIN users ON files.user_id = users.telegram_id
            WHERE files.declaration_number = ?
        """, (at,))
        u = cursor.fetchone()
        owner_name = u[0] if u and u[0] else (f"@{u[1]}" if u and u[1] else (str(u[2]) if u else "Noma'lum"))
        dup_text = f"⚠️ **Diqqat!** Bu fayl avval kiritilgan.\n📄 AT: `{at}`\n👤 Kiritgan: **{owner_name}**"
        await bot.edit_message_text(dup_text, chat_id=message.from_user.id, message_id=msg.message_id, parse_mode="Markdown")

    except Exception as e:
        error_text = f"❌ **Xatolik yuz berdi:** {str(e)}"
        await bot.edit_message_text(error_text, chat_id=message.from_user.id, message_id=msg.message_id)

    finally:
        if os.path.exists(path): os.remove(path)

# ================= CALLBACKS =================
@dp.callback_query(F.data == "back_to_main")
async def back_to_main_h(cb: CallbackQuery):
    await cb.message.delete()
    role = get_user_role(cb.from_user.id)
    if is_admin(role):
        await bot.send_message(cb.from_user.id, "🏢 **Boshqaruv paneli**", reply_markup=main_admin_keyboard(), parse_mode="Markdown")
    else:
        await bot.send_message(cb.from_user.id, "📝 **PDF fayllarni yuboring.**", reply_markup=sub_admin_keyboard(), parse_mode="Markdown")

@dp.message(F.text == "👥 Foydalanuvchilar boshqaruvi")
async def user_management(message: types.Message):
    if is_admin(get_user_role(message.from_user.id)):
        await message.answer("👥 **Foydalanuvchilar ro'yxati:**", reply_markup=users_list_inline(), parse_mode="Markdown")

@dp.callback_query(F.data == "back_to_users")
async def back_to_u(cb: CallbackQuery):
    await cb.message.edit_text("👥 **Foydalanuvchilar ro'yxati:**", reply_markup=users_list_inline(), parse_mode="Markdown")

@dp.callback_query(F.data.startswith("show_id:"))
async def show_user_id(cb: CallbackQuery):
    tg_id = cb.data.split(":")[1]
    await cb.answer(f"Telegram ID: {tg_id}", show_alert=True)

@dp.callback_query(F.data.startswith("u_info:"))
async def user_info_detail(cb: CallbackQuery):
    tg_id = cb.data.split(":")[1]
    cursor.execute("SELECT full_name, role, username FROM users WHERE telegram_id=?", (tg_id,))
    u = cursor.fetchone()
    if not u: return await cb.answer("Foydalanuvchi topilmadi")

    cursor.execute("SELECT COUNT(*), SUM(bojxona_sum) FROM files WHERE user_id=?", (tg_id,))
    stat = cursor.fetchone()

    d_name = u[0] if u[0] else (f"@{u[2]}" if u[2] else str(tg_id))

    text = (f"👤 **Foydalanuvchi:** {d_name}\n"
            f"🎭 **Rol:** `{u[1].upper()}`\n\n"
            f"📊 **Jami yuklangan:** {stat[0]} ta\n"
            f"💰 **Jami summa:** {format_currency(stat[1])}")

    role = get_user_role(cb.from_user.id)
    kb = InlineKeyboardMarkup(inline_keyboard=[])
    role_buttons = []
    if role == "super_admin":
        role_buttons.append(InlineKeyboardButton(text="👑 Chief-Admin", callback_data=f"role:chief_admin:{tg_id}"))
    role_buttons += [
        InlineKeyboardButton(text="⭐ Sub-Admin", callback_data=f"role:sub_admin:{tg_id}"),
        InlineKeyboardButton(text="👤 User", callback_data=f"role:user:{tg_id}")
    ]
    kb.inline_keyboard.append(role_buttons)
    kb.inline_keyboard.append([InlineKeyboardButton(text="🧹 Fayllarni boshqarish", callback_data=f"u_clear_files:{tg_id}")])
    kb.inline_keyboard.append([InlineKeyboardButton(text="⬅️ Orqaga", callback_data="back_to_users")])
    await cb.message.edit_text(text, reply_markup=kb, parse_mode="Markdown")

@dp.callback_query(F.data.startswith("role:"))
async def change_role(cb: CallbackQuery):
    _, new_role, tg_id = cb.data.split(":")
    if get_user_role(cb.from_user.id) != "super_admin" and new_role == "chief_admin":
        return await cb.answer("Faqat super admin chief-admin bera oladi!")
    cursor.execute("UPDATE users SET role=? WHERE telegram_id=?", (new_role, tg_id))
    conn.commit()
    try:
        if new_role == "chief_admin":
            await bot.send_message(tg_id, "✅ **Siz chief-admin bo'ldingiz!**", reply_markup=main_admin_keyboard(), parse_mode="Markdown")
        elif new_role == "sub_admin":
            await bot.send_message(tg_id, "✅ **Siz sub-admin bo'ldingiz!**", reply_markup=sub_admin_keyboard(), parse_mode="Markdown")
        else:
            await bot.send_message(tg_id, "⚠️ **Sizning kirish huquqingiz bekor qilindi.**", reply_markup=types.ReplyKeyboardRemove())
    except Exception as e:
        logging.error(f"Rul yuborish xatosi: {e}")
    await cb.answer("✅ Rol o'zgartirildi!")
    await user_info_detail(cb)

@dp.callback_query(F.data.startswith("confirm_u_del:"))
async def confirm_u_del(cb: CallbackQuery):
    tg_id = cb.data.split(":")[1]
    cursor.execute("SELECT telegram_id, username, full_name, role FROM users WHERE telegram_id != ?", (SUPER_ADMIN_ID,))
    users = cursor.fetchall()
    kb = InlineKeyboardMarkup(inline_keyboard=[])
    for t_id, user, name, role in users:
        display_name = name if name else (f"@{user}" if user else str(t_id))
        role_status = "👑" if role == "chief_admin" else ("⭐" if role == "sub_admin" else "👤")
        if str(t_id) == tg_id:
            kb.inline_keyboard.append([
                InlineKeyboardButton(text="🆔", callback_data=f"show_id:{t_id}"),
                InlineKeyboardButton(text=f"{role_status} {display_name[:15]}", callback_data=f"u_info:{t_id}"),
                InlineKeyboardButton(text="✅ Tasdiqla", callback_data=f"do_u_del:{t_id}"),
                InlineKeyboardButton(text="❌ Bekor", callback_data=f"cancel_u_del:{t_id}")
            ])
        else:
            kb.inline_keyboard.append([
                InlineKeyboardButton(text="🆔", callback_data=f"show_id:{t_id}"),
                InlineKeyboardButton(text=f"{role_status} {display_name[:15]}", callback_data=f"u_info:{t_id}"),
                InlineKeyboardButton(text="🗑️", callback_data=f"confirm_u_del:{t_id}")
            ])
    kb.inline_keyboard.append([InlineKeyboardButton(text="⬅️ Asosiy menyuga qaytish", callback_data="back_to_main")])
    await cb.message.edit_reply_markup(reply_markup=kb)

@dp.callback_query(F.data.startswith("do_u_del:"))
async def do_u_del(cb: CallbackQuery):
    tg_id = cb.data.split(":")[1]
    cursor.execute("DELETE FROM users WHERE telegram_id=?", (tg_id,))
    conn.commit()
    await cb.answer("✅ Foydalanuvchi o'chirildi")
    await cb.message.edit_reply_markup(reply_markup=users_list_inline())

@dp.callback_query(F.data.startswith("cancel_u_del:"))
async def cancel_u_del(cb: CallbackQuery):
    tg_id = cb.data.split(":")[1]
    await cb.answer("❌ Bekor qilindi")
    await cb.message.edit_reply_markup(reply_markup=users_list_inline())

@dp.message(F.text == "📊 Umumiy statistika")
async def total_stats_start(message: types.Message):
    if not is_admin(get_user_role(message.from_user.id)): return
    now = datetime.datetime.now()
    await message.answer("📅 Statistika uchun **BOSHLANISH** sanasini tanlang:",
                         reply_markup=get_calendar_markup(now.year, now.month, "all_stat", "none", "start"), parse_mode="Markdown")

@dp.message(F.text == "📥 Excel Hisobot")
async def excel_report_start(message: types.Message):
    if not is_admin(get_user_role(message.from_user.id)): return
    now = datetime.datetime.now()
    await message.answer("📥 Excel uchun **BOSHLANISH** sanasini tanlang:",
                         reply_markup=get_calendar_markup(now.year, now.month, "excel", "none", "start"), parse_mode="Markdown")

@dp.message(F.text.startswith("📍"))
async def post_selected(message: types.Message):
    if not is_admin(get_user_role(message.from_user.id)): return
    post_name = message.text.replace("📍 ", "")
    now = datetime.datetime.now()
    await message.answer(f"🏢 **{post_name}** uchun **BOSHLANISH** sanasini tanlang:",
                         reply_markup=get_calendar_markup(now.year, now.month, "post", post_name, "start"), parse_mode="Markdown")

@dp.callback_query(F.data.startswith(("nav:", "cal:")))
async def process_calendar_logic(cb: CallbackQuery):
    data = cb.data.split(":")
    cmd, t_type, t_id, side, year, month = data[0], data[1], data[2], data[3], int(data[4]), int(data[5])

    if cmd == "nav":
        action = data[6]
        month = month - 1 if action == "prev" else month + 1
        if month < 1: month = 12; year -= 1
        elif month > 12: month = 1; year += 1
        await cb.message.edit_reply_markup(reply_markup=get_calendar_markup(year, month, t_type, t_id, side))

    elif cmd == "cal":
        day = int(data[6])
        selected_date = f"{year}-{month:02d}-{day:02d}"
        if side == "start":
            user_sessions[cb.from_user.id] = {"start": selected_date}
            await cb.message.edit_text(f"📍 Boshlanish: `{selected_date}`\n📅 **YAKUNLANISH** sanasini tanlang:",
                                       reply_markup=get_calendar_markup(year, month, t_type, t_id, "end"), parse_mode="Markdown")
        else:
            start_date = user_sessions.get(cb.from_user.id, {}).get("start")
            end_date = selected_date

            if t_type == "all_stat":
                cursor.execute("SELECT kirish_posti, COUNT(id), SUM(bojxona_sum) FROM files WHERE date_only BETWEEN ? AND ? GROUP BY kirish_posti", (start_date, end_date))
                stats = cursor.fetchall()
                if not stats:
                    await cb.message.edit_text("❌ Bu oraliqda ma'lumot yo'q.", reply_markup=InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="⬅️ Orqaga", callback_data="back_to_main")]]))
                    return

                labels = [str(s[0]) for s in stats]
                values = [int(s[2] or 0) for s in stats]
                total_sum = sum(values)

                def autopct_format(pct):
                    absolute = int(round(pct / 100. * total_sum))
                    return f"{pct:.1f}%\n{format_currency(absolute)}"

                plt.figure(figsize=(12, 10), dpi=120)
                colors = cm.viridis([i / len(values) for i in range(len(values))])
                wedges, texts, autotexts = plt.pie(
                    values,
                    labels=labels,
                    autopct=autopct_format,
                    startangle=140,
                    colors=colors,
                    shadow=True,
                    wedgeprops=dict(width=0.3, edgecolor='w'),
                    pctdistance=0.80,
                    labeldistance=1.15
                )
                plt.setp(autotexts, size=11, weight="bold", color="white")
                plt.setp(texts, size=13)
                plt.title(f"Postlar bo'yicha tushum taqsimoti\n({start_date} - {end_date})", fontsize=16, fontweight='bold')

                centre_circle = plt.Circle((0,0),0.70,fc='white')
                fig = plt.gcf()
                fig.gca().add_artist(centre_circle)

                plt.legend(wedges, labels, title="Postlar", loc="center left", bbox_to_anchor=(1, 0, 0.5, 1), fontsize=10)

                buf_pie = io.BytesIO()
                plt.savefig(buf_pie, format='png', bbox_inches='tight')
                buf_pie.seek(0)
                plt.close()

                res = f"📊 **Umumiy Statistika**\n📅 `{start_date}` dan `{end_date}` gacha\n\n"
                t_cnt, t_sum = 0, 0
                for p, c, s in stats:
                    res += f"🏢 **{p}**: {c} ta | {format_currency(s)}\n"
                    t_cnt += c; t_sum += (s or 0)
                res += f"\n🔥 **JAMI:** {t_cnt} ta\n💰 **SUMMA:** {format_currency(t_sum)}"

                await cb.message.delete()
                photo_pie = types.BufferedInputFile(buf_pie.read(), filename="stat_pie.png")
                await bot.send_photo(cb.from_user.id, photo=photo_pie, caption=res, parse_mode="Markdown")

                cursor.execute("""
                    SELECT strftime('%Y-%m', date_only) as month, kirish_posti, SUM(bojxona_sum) 
                    FROM files 
                    WHERE date_only BETWEEN ? AND ? 
                    GROUP BY month, kirish_posti 
                    ORDER BY month
                """, (start_date, end_date))
                monthly_stats = cursor.fetchall()
                if monthly_stats:
                    df = pd.DataFrame(monthly_stats, columns=['month', 'posti', 'summa'])
                    df['month'] = pd.to_datetime(df['month'] + '-01')
                    posts = df['posti'].unique()
                    months = df['month'].unique()

                    plt.figure(figsize=(12, 8), dpi=100)
                    bar_width = 0.15
                    for i, post in enumerate(posts):
                        post_data = df[df['posti'] == post]
                        plt.bar(post_data['month'] + pd.Timedelta(days=i*2), post_data['summa'], width=bar_width, label=post, color=colors[i % len(colors)])

                    plt.title(f"Oylar bo'yicha postlar taqoslashi\n({start_date} - {end_date})", fontsize=16, fontweight='bold')
                    plt.xlabel("Oylar", fontsize=14)
                    plt.ylabel("Summa (so'm)", fontsize=14)
                    plt.xticks(months, [m.strftime('%Y-%m') for m in months], rotation=45)
                    plt.legend(title="Postlar", bbox_to_anchor=(1.05, 1), loc='upper left')
                    plt.grid(axis='y', linestyle='--', alpha=0.7)
                    plt.tight_layout()

                    buf_bar = io.BytesIO()
                    plt.savefig(buf_bar, format='png', bbox_inches='tight')
                    buf_bar.seek(0)
                    plt.close()

                    photo_bar = types.BufferedInputFile(buf_bar.read(), filename="stat_bar.png")
                    await bot.send_photo(cb.from_user.id, photo=photo_bar, caption="📈 **Oylar bo'yicha taqoslash**", parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="⬅️ Orqaga", callback_data="back_to_main")]]))

                else:
                    await bot.send_message(cb.from_user.id, "❌ Oylar bo'yicha ma'lumot yo'q.", reply_markup=InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="⬅️ Orqaga", callback_data="back_to_main")]]))

            elif t_type == "excel":
                await cb.message.edit_text("⏳ Excel fayl tayyorlanmoqda...")
                query = """
                    SELECT f.timestamp, f.declaration_number, f.deklarant_fish, f.avto, f.tirkama, f.kirish_posti, f.bojxona_sum, u.full_name
                    FROM files f JOIN users u ON f.user_id = u.telegram_id
                    WHERE f.date_only BETWEEN ? AND ?
                """
                df_details = pd.read_sql_query(query, conn, params=(start_date, end_date))
                df_details.columns = ['Vaqt', 'AT Raqam', 'Deklarant', 'Avto', 'Tirkama', 'Post', 'Summa', 'Kiritdi']

                cursor.execute("SELECT kirish_posti, COUNT(id), SUM(bojxona_sum) FROM files WHERE date_only BETWEEN ? AND ? GROUP BY kirish_posti", (start_date, end_date))
                df_post = pd.DataFrame(cursor.fetchall(), columns=['Post', 'Soni', 'Summa'])

                cursor.execute("""
                    SELECT strftime('%Y-%m', date_only) as OY, kirish_posti, SUM(bojxona_sum) 
                    FROM files 
                    WHERE date_only BETWEEN ? AND ? 
                    GROUP BY OY, kirish_posti 
                    ORDER BY OY
                """, (start_date, end_date))
                df_monthly = pd.DataFrame(cursor.fetchall(), columns=['OY', 'Post', 'Summa'])
                df_pivot = df_monthly.pivot(index='Post', columns='OY', values='Summa').fillna(0)
                df_pivot['Jami'] = df_pivot.sum(axis=1)
                df_pivot.loc['Jami'] = df_pivot.sum()

                df_growth = df_pivot.copy()
                for col in df_growth.columns[1:-1]:
                    prev_col = df_growth.columns[df_growth.columns.get_loc(col) - 1]
                    df_growth[col] = ((df_growth[col] - df_growth[prev_col]) / df_growth[prev_col]) * 100 if df_growth[prev_col] != 0 else 0

                file_path = f"Hisobot_{start_date}_{end_date}.xlsx"
                wb = Workbook()
                ws_details = wb.active
                ws_details.title = "Tafsilotlar"
                for r in pd.DataFrame([df_details.columns]).itertuples(index=False):
                    ws_details.append(r)
                for r in df_details.itertuples(index=False):
                    ws_details.append(r)

                ws_post = wb.create_sheet("Postlar bo'yicha")
                for r in pd.DataFrame([df_post.columns]).itertuples(index=False):
                    ws_post.append(r)
                for r in df_post.itertuples(index=False):
                    ws_post.append(r)

                chart_post = BarChart()
                chart_post.type = "col"
                chart_post.style = 10
                chart_post.title = "Postlar bo'yicha summa"
                chart_post.y_axis.title = "Summa"
                chart_post.x_axis.title = "Post"
                data = Reference(ws_post, min_col=3, min_row=1, max_row=len(df_post)+1, max_col=3)
                cats = Reference(ws_post, min_col=1, min_row=2, max_row=len(df_post)+1)
                chart_post.add_data(data, titles_from_data=True)
                chart_post.set_categories(cats)
                chart_post.shape = 4
                ws_post.add_chart(chart_post, "E2")

                ws_monthly = wb.create_sheet("Oylar bo'yicha")
                for r in pd.DataFrame([df_pivot.columns]).itertuples(index=False):
                    ws_monthly.append(r)
                for index, row in df_pivot.iterrows():
                    ws_monthly.append(list(row))
                ws_monthly.append([''] * len(df_pivot.columns))
                ws_monthly.append([''] + list(df_growth.columns[1:]))
                for index, row in df_growth.iterrows():
                    row_list = [index] + [round(val, 2) if isinstance(val, (int, float)) else val for val in row[1:]]
                    ws_monthly.append(row_list)

                green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                red_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
                bold_font = Font(bold=True)
                center_align = Alignment(horizontal='center')
                border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                growth_start_row = len(df_pivot) + 3

                for row in ws_monthly.iter_rows(min_row=1, max_row=ws_monthly.max_row, min_col=1, max_col=ws_monthly.max_column):
                    for cell in row:
                        cell.alignment = center_align
                        cell.border = border
                        if cell.row in (1, growth_start_row):
                            cell.font = bold_font
                        if cell.row >= growth_start_row + 1 and cell.column > 1 and cell.value is not None:
                            try:
                                val = float(cell.value)
                                if val > 0:
                                    cell.fill = green_fill
                                elif val < 0:
                                    cell.fill = red_fill
                                cell.number_format = '0.00%'
                            except (ValueError, TypeError):
                                pass

                chart_monthly = BarChart()
                chart_monthly.type = "col"
                chart_monthly.style = 12
                chart_monthly.title = "Oylar bo'yicha summa"
                chart_monthly.y_axis.title = "Summa"
                chart_monthly.x_axis.title = "Post"
                data_monthly = Reference(ws_monthly, min_col=2, min_row=2, max_row=len(df_pivot)+1, max_col=len(df_pivot.columns))
                cats_monthly = Reference(ws_monthly, min_col=1, min_row=3, max_row=len(df_pivot)+1)
                chart_monthly.add_data(data_monthly, titles_from_data=True)
                chart_monthly.set_categories(cats_monthly)
                chart_monthly.dataLabels = DataLabelList()
                chart_monthly.dataLabels.showVal = True
                ws_monthly.add_chart(chart_monthly, "A{}".format(len(df_pivot) + len(df_growth) + 5))

                wb.save(file_path)

                await cb.message.delete()
                await bot.send_document(cb.from_user.id, FSInputFile(file_path), caption=f"📅 {start_date} - {end_date}\n📊 Barcha yuklangan ma'lumotlar")
                if os.path.exists(file_path): os.remove(file_path)

            elif t_type == "post":
                cursor.execute("""SELECT users.full_name, users.username, COUNT(files.id), SUM(files.bojxona_sum), users.telegram_id
                                FROM users JOIN files ON users.telegram_id = files.user_id
                                WHERE files.kirish_posti LIKE ? AND files.date_only BETWEEN ? AND ? GROUP BY users.telegram_id""", (f"%{t_id}%", start_date, end_date))
                u_stats = cursor.fetchall()
                kb = InlineKeyboardMarkup(inline_keyboard=[])
                for name, nick, cnt, sm, uid in u_stats:
                    kb.inline_keyboard.append([InlineKeyboardButton(text=f"👤 {name or nick} | {cnt} ta | {format_currency(sm)}", callback_data=f"u_clear_files:{uid}")])
                kb.inline_keyboard.append([InlineKeyboardButton(text="⬅️ Orqaga", callback_data="back_to_main")])
                await cb.message.edit_text(f"🏢 **Post:** {t_id}\n📅 `{start_date}` / `{end_date}`", reply_markup=kb, parse_mode="Markdown")

            elif t_type == "my_report":
                cursor.execute("SELECT COUNT(*), SUM(bojxona_sum) FROM files WHERE user_id=? AND date_only BETWEEN ? AND ?", (cb.from_user.id, start_date, end_date))
                c, s = cursor.fetchone()
                text = f"📊 **Hisobotingiz**\n📅 `{start_date}` dan `{end_date}` gacha\n\n📄 **Jami:** {c} ta\n💰 **Summa:** {format_currency(s)}"
                kb = InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="🧹 Fayllarni boshqarish", callback_data=f"my_clear_files:{cb.from_user.id}:{start_date}:{end_date}")]])
                kb.inline_keyboard.append([InlineKeyboardButton(text="⬅️ Orqaga", callback_data="back_to_main")])
                await cb.message.edit_text(text, reply_markup=kb, parse_mode="Markdown")

@dp.message(F.text == "📊 Mening statistikam")
async def my_report_init_msg(message: types.Message):
    now = datetime.datetime.now()
    await message.answer("📅 Hisobot uchun **BOSHLANISH** sanasini tanlang:",
                         reply_markup=get_calendar_markup(now.year, now.month, "my_report", "none", "start"), parse_mode="Markdown")

def get_user_files_kb(uid, start_date=None, end_date=None, is_sub=False):
    if start_date and end_date:
        cursor.execute("SELECT id, declaration_number, avto FROM files WHERE user_id=? AND date_only BETWEEN ? AND ? ORDER BY id DESC LIMIT 20", (uid, start_date, end_date))
    else:
        cursor.execute("SELECT id, declaration_number, avto FROM files WHERE user_id=? ORDER BY id DESC LIMIT 20", (uid,))
    files = cursor.fetchall()
    kb = InlineKeyboardMarkup(inline_keyboard=[])
    for fid, at, avto in files:
        kb.inline_keyboard.append([
            InlineKeyboardButton(text=f"🚛 {avto} ({at})", callback_data=f"car_info:{at}"),
            InlineKeyboardButton(text="🗑️", callback_data=f"confirm_del:{fid}:{uid}:{start_date if start_date else ''}:{end_date if end_date else ''}:{'sub' if is_sub else 'admin'}")
        ])
    kb.inline_keyboard.append([InlineKeyboardButton(text="⬅️ Orqaga", callback_data=f"{'my_report_back' if is_sub else 'u_info'}:{uid}:{start_date if start_date else ''}:{end_date if end_date else ''}")])
    return kb

@dp.callback_query(F.data.startswith("my_clear_files:"))
async def sub_clear_files(cb: CallbackQuery):
    _, uid, start_date, end_date = cb.data.split(":")
    if int(uid) != cb.from_user.id or get_user_role(cb.from_user.id) != "sub_admin": return
    files_text = "🔍 **Oxirgi fayllaringiz:**"
    await cb.message.edit_text(files_text, reply_markup=get_user_files_kb(uid, start_date, end_date, is_sub=True), parse_mode="Markdown")

@dp.callback_query(F.data.startswith("my_report_back:"))
async def my_report_back(cb: CallbackQuery):
    data = cb.data.split(":")
    uid = data[1] if len(data) > 1 else cb.from_user.id
    start_date = data[2] if len(data) > 2 else None
    end_date = data[3] if len(data) > 3 else None
    cursor.execute("SELECT COUNT(*), SUM(bojxona_sum) FROM files WHERE user_id=? AND date_only BETWEEN ? AND ?", (cb.from_user.id, start_date, end_date))
    c, s = cursor.fetchone()
    text = f"📊 **Hisobotingiz**\n📅 `{start_date}` dan `{end_date}` gacha\n\n📄 **Jami:** {c} ta\n💰 **Summa:** {format_currency(s)}"
    kb = InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="🧹 Fayllarni boshqarish", callback_data=f"my_clear_files:{cb.from_user.id}:{start_date}:{end_date}")]])
    kb.inline_keyboard.append([InlineKeyboardButton(text="⬅️ Orqaga", callback_data="back_to_main")])
    await cb.message.edit_text(text, reply_markup=kb, parse_mode="Markdown")

@dp.callback_query(F.data.startswith("u_clear_files:"))
async def admin_clear_user_files(cb: CallbackQuery):
    uid = cb.data.split(":")[1]
    if not is_admin(get_user_role(cb.from_user.id)): return
    files_text = "🔍 **Foydalanuvchining oxirgi fayllari:**"
    await cb.message.edit_text(files_text, reply_markup=get_user_files_kb(uid), parse_mode="Markdown")

@dp.callback_query(F.data.startswith("confirm_del:"))
async def confirm_del(cb: CallbackQuery):
    data = cb.data.split(":")
    fid, uid = data[1], data[2]
    start_date = data[3] if len(data) > 3 else ''
    end_date = data[4] if len(data) > 4 else ''
    mode = data[5] if len(data) > 5 else 'admin'
    is_sub = mode == 'sub'
    if is_sub and int(uid) != cb.from_user.id: return
    if start_date == '': start_date = None
    if end_date == '': end_date = None
    cursor.execute("SELECT id, declaration_number, avto FROM files WHERE user_id=? ORDER BY id DESC LIMIT 20", (uid,))
    files = cursor.fetchall()
    kb = InlineKeyboardMarkup(inline_keyboard=[])
    for f_id, at, avto in files:
        if f_id == int(fid):
            kb.inline_keyboard.append([
                InlineKeyboardButton(text=f"🚛 {avto} ({at})", callback_data=f"car_info:{at}"),
                InlineKeyboardButton(text="✅ Tasdiqla", callback_data=f"do_del:{fid}:{uid}:{start_date if start_date else ''}:{end_date if end_date else ''}:{mode}"),
                InlineKeyboardButton(text="❌ Bekor", callback_data=f"cancel_del:{fid}:{uid}:{start_date if start_date else ''}:{end_date if end_date else ''}:{mode}")
            ])
        else:
            kb.inline_keyboard.append([
                InlineKeyboardButton(text=f"🚛 {avto} ({at})", callback_data=f"car_info:{at}"),
                InlineKeyboardButton(text="🗑️", callback_data=f"confirm_del:{f_id}:{uid}:{start_date if start_date else ''}:{end_date if end_date else ''}:{mode}")
            ])
    kb.inline_keyboard.append([InlineKeyboardButton(text="⬅️ Orqaga", callback_data=f"{'my_report_back' if is_sub else 'u_info'}:{uid}:{start_date if start_date else ''}:{end_date if end_date else ''}")])
    await cb.message.edit_reply_markup(reply_markup=kb)

@dp.callback_query(F.data.startswith("do_del:"))
async def do_del(cb: CallbackQuery):
    data = cb.data.split(":")
    fid, uid = data[1], data[2]
    start_date = data[3] if len(data) > 3 else ''
    end_date = data[4] if len(data) > 4 else ''
    mode = data[5] if len(data) > 5 else 'admin'
    if mode == 'sub' and int(uid) != cb.from_user.id: return
    if start_date == '': start_date = None
    if end_date == '': end_date = None
    cursor.execute("DELETE FROM files WHERE id=?", (fid,))
    conn.commit()
    await cb.answer("✅ Fayl o'chirildi")
    await cb.message.edit_reply_markup(reply_markup=get_user_files_kb(uid, start_date, end_date, is_sub=(mode == 'sub')))

@dp.callback_query(F.data.startswith("cancel_del:"))
async def cancel_del(cb: CallbackQuery):
    data = cb.data.split(":")
    fid, uid = data[1], data[2]
    start_date = data[3] if len(data) > 3 else ''
    end_date = data[4] if len(data) > 4 else ''
    mode = data[5] if len(data) > 5 else 'admin'
    if start_date == '': start_date = None
    if end_date == '': end_date = None
    await cb.answer("❌ Bekor qilindi")
    await cb.message.edit_reply_markup(reply_markup=get_user_files_kb(uid, start_date, end_date, is_sub=(mode == 'sub')))

@dp.callback_query(F.data.startswith("car_info:"))
async def car_info_callback(cb: CallbackQuery):
    at = cb.data.split(":")[1]
    cursor.execute("SELECT * FROM files WHERE declaration_number=?", (at,))
    f = cursor.fetchone()
    if f:
        res = f"📑 Ma'lumot:\nAT: {f[2]}\nSumma: {format_currency(f[3])}\nDeklarant: {f[4]}\nAvto: {f[5]} / {f[6]}\nPost: {f[7]}\nVaqt: {f[8]}"
        await cb.answer(res, show_alert=True)

async def main():
    logging.info("Bot Render.com da ishga tushdi")
    await dp.start_polling(bot)

if __name__ == "__main__":
    try:
        asyncio.run(main())
    except (KeyboardInterrupt, SystemExit):
        logging.error("Bot to'xtatildi!")
    except Exception as e:
        logging.error(f"Bot ishida katta xato: {e}", exc_info=True) 